"""Give each rater one self-contained folder whose files play in sheet order.

The first packaging put every clip in one `audio/` directory named by a hash and
listed the same hashes, shuffled, in each rater's sheet. A rater therefore had
to search the folder for each token in turn -- and the three sheets are
deliberately in *different* orders, so no single naming of one shared folder can
match all of them.

So each rater gets their own folder: their sheet, and the clips numbered 001,
002, ... in the order that sheet asks for them. Play them in order, fill the row
with the same number, done.

Numbering does not leak anything. The order was already randomised per rater
before this ran, so a position says nothing about which system produced the clip
-- and the repeated clips, which measure whether a rater agrees with themselves,
now appear under two different numbers instead of the same visible hash twice.
That makes the repeat check stronger than it was.

`audio/` stays as the source the packs are rebuilt from; only the per-rater
folders go to the raters.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
from pathlib import Path
from typing import Any

# Column widths, a frozen header and a dropdown are things a CSV cannot carry,
# and the people filling these in open them in a spreadsheet. The CSV stays
# beside it as the format the scorer reads and as the fallback.
COLUMN_WIDTHS = {"行号": 6, "音频文件": 12, "emotion_heard": 16, "mos_1_to_5": 12, "defect": 8}
CHOICES = "中立,开心,生气,吃惊,难过,说不上来"


def write_workbook(
    path: Path, header: list[str], rows: list[dict[str, Any]], answers: list[str]
) -> None:
    """Same sheet as the CSV, with the widths and the dropdown filled in."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("  （没有 openpyxl，只出 csv）")
        return

    book = Workbook()
    sheet = book.active
    sheet.title = path.stem
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for line, row in enumerate(rows, start=1):
        sheet.append([line, f"{row['position']}.wav", *([""] * len(answers))])

    for index, name in enumerate(header, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(name, 40)
        sheet.cell(row=1, column=index).alignment = Alignment(horizontal="center")
    # The header scrolls off after twenty rows otherwise, and a rater who cannot
    # see which column is which starts guessing.
    sheet.freeze_panes = "A2"

    if "emotion_heard" in header:
        column = sheet.cell(row=1, column=header.index("emotion_heard") + 1).column_letter
        rule = DataValidation(type="list", formula1=f'"{CHOICES}"', allow_blank=True)
        rule.error = "只能填这六个之一：" + CHOICES.replace(",", " / ")
        sheet.add_data_validation(rule)
        rule.add(f"{column}2:{column}{len(rows) + 1}")
    if "mos_1_to_5" in header:
        column = sheet.cell(row=1, column=header.index("mos_1_to_5") + 1).column_letter
        rule = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
        rule.error = "只能填 1 到 5 的整数"
        sheet.add_data_validation(rule)
        rule.add(f"{column}2:{column}{len(rows) + 1}")

    book.save(str(path))


def answer_columns(pack: Path, rater: str) -> list[str]:
    """The columns a rater fills, recovered from a folder written by an earlier run."""
    existing = pack / f"rater{rater}" / f"rater{rater}.csv"
    if not existing.is_file():
        return ["mos_1_to_5", "defect", "note"]
    with existing.open(encoding="utf-8-sig", newline="") as handle:
        fields = list(csv.DictReader(handle).fieldnames or [])
    return [name for name in fields if name not in ("行号", "音频文件", "序号", "token")]


def read_sheet(path: Path) -> tuple[list[str], list[str]]:
    """Column names and the token of each row, in sheet order."""
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fields = list(reader.fieldnames or [])
        tokens = [row[fields[0]].split("#")[0] for row in reader]
    return fields, tokens


def repack(pack: Path, readme: str) -> dict[str, Any]:
    key: dict[str, Any] = json.loads((pack / "key.json").read_text(encoding="utf-8"))
    by_token = {entry["token"]: entry for entry in key["entries"]}
    source = pack / "audio"
    if not source.is_dir():
        raise SystemExit(f"{source} 不在——重打包需要原始音频")

    # A pack that has already been repacked has no sheet at the root; its per
    # rater order lives in the key instead. Reading it back from there makes
    # this rerunnable, which matters because the layout has now changed twice.
    previous = key.get("raters") or {}
    sheets = sorted(pack.glob("rater*.csv"))
    raters = [s.stem.replace("rater", "") for s in sheets] or sorted(previous)

    layout: dict[str, list[dict[str, Any]]] = {}
    for rater in raters:
        sheet = pack / f"rater{rater}.csv"
        if sheet.is_file():
            fields, tokens = read_sheet(sheet)
        else:
            fields = ["位置", *answer_columns(pack, rater)]
            tokens = [row["token"] for row in previous[rater]]
        folder = pack / f"rater{rater}"
        if folder.exists():
            shutil.rmtree(folder)
        folder.mkdir(parents=True)

        rows = []
        for position, token in enumerate(tokens, start=1):
            number = f"{position:03d}"
            clip = source / f"{token}.wav"
            if not clip.is_file():
                raise SystemExit(f"{clip} 不在，无法重打包 {sheet.name}")
            shutil.copy(clip, folder / f"{number}.wav")
            rows.append({"position": number, "token": token, **by_token.get(token, {})})

        answers = fields[1:]
        header = ["行号", "音频文件", *answers]
        with (folder / sheet.name).open("w", encoding="utf-8-sig", newline="") as sink:
            writer = csv.writer(sink)
            writer.writerow(header)
            for line, row in enumerate(rows, start=1):
                writer.writerow([line, f"{row['position']}.wav", *([""] * len(answers))])
        write_workbook(folder / f"{sheet.stem}.xlsx", header, rows, answers)
        layout[rater] = rows
        print(f"  rater{rater}: {len(rows)} 行 -> {folder.name}/")
        # The sheet at the pack root would now disagree with the numbered copy.
        # A rater with it open in a spreadsheet holds a lock on Windows, and
        # failing the whole repack over a stale file nobody should be using is
        # the wrong trade -- warn and carry on.
        if sheet.is_file():
            try:
                sheet.unlink()
            except OSError:
                print(f"  ⚠ 删不掉 {sheet.name}（有程序开着它）——**那是旧表，别再填它**")

    key["raters"] = layout
    key["layout"] = (
        "each rater has their own folder; clips are numbered in that rater's sheet order"
    )
    (pack / "key.json").write_text(
        json.dumps(key, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    (pack / "README.txt").write_text(readme, encoding="utf-8")
    return key


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pack", type=Path, required=True)
    parser.add_argument("--readme", type=Path, required=True, help="the scoring guide to install")
    args = parser.parse_args()

    key = repack(args.pack, args.readme.read_text(encoding="utf-8"))
    print(f"\n{args.pack}: {len(key['raters'])} 位评分员，各自一个文件夹")
    print("**key.json 评分员不要看**")


if __name__ == "__main__":
    main()
